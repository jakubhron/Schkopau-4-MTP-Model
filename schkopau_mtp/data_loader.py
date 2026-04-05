"""
Data loading, cleaning, and feature engineering for the Schkopau MTP model.

Responsibilities:
  - Read the input Excel file (tabs Block_A, Block_B)
  - Merge common + block-specific columns into one DataFrame
  - Filter to the configured date range
  - Compute derived columns (seasons, shutdown days, cost curve per block, etc.)
"""

from __future__ import annotations

from typing import List, Tuple

import os

import numpy as np
import pandas as pd

from . import config as cfg


# ====================================================================
#  PUBLIC API
# ====================================================================


def load_and_prepare() -> Tuple[pd.DataFrame, dict]:
    """
    Load input data from Block_A / Block_B tabs, merge, clean, and compute
    all derived columns.

    Returns
    -------
    df : pd.DataFrame
        Hourly time-series with all columns needed by the optimisation model.
        Block-specific columns are suffixed ``_A``, ``_B``.
    meta : dict
        Auxiliary data consumed downstream (per-block Pmax_eff, etc.).
    """
    df = _read_input()
    df = _filter_date_range(df)
    df = _add_time_index(df)
    df = _add_gridfee(df)
    df = _add_season_and_dow(df)
    df, cost_meta = _compute_cost_curves(df)
    df = _compute_coal_curves(df)
    starts_data = _read_starts_tab()
    coal_limits = _read_coal_constrains_tab() if cfg.USE_COAL_CONSTRAINS else {}

    meta = {**cost_meta, "starts": starts_data, "coal_limits": coal_limits}
    return df, meta


# ====================================================================
#  INTERNAL HELPERS
# ====================================================================

# Columns treated as common (identical across tabs).  Everything else is block-specific.
_COMMON_COLS = {
    "date", "hour", "month", "price", "eua", "coal price",
    "grid fee", "warme", "weekday",
}


def _is_block_col(col_name: str) -> bool:
    return col_name.strip().lower() not in _COMMON_COLS


def _read_input() -> pd.DataFrame:
    """Read Block_A and Block_B tabs and merge into a single DataFrame."""
    import shutil, tempfile
    # Copy to a temp file so we can read even when Excel has the file open
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    shutil.copy2(cfg.INPUT_FILE, tmp_path)
    try:
        dfs: dict[str, pd.DataFrame] = {}
        for block in cfg.BLOCKS:
            sheet = f"Block_{block}"
            dfs[block] = pd.read_excel(tmp_path, sheet_name=sheet)
    finally:
        os.unlink(tmp_path)

    # Use Block_A as base for common columns
    base = dfs[cfg.BLOCKS[0]].copy()
    common_cols = [c for c in base.columns if not _is_block_col(c)]
    df = base[common_cols].copy()

    # Merge block-specific columns from each tab
    for block in cfg.BLOCKS:
        block_df = dfs[block]
        for col in block_df.columns:
            if _is_block_col(col):
                df[f"{col}_{block}"] = block_df[col].values

    return df


def _filter_date_range(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only rows within the configured START_DATE..END_DATE window."""
    df = df[df["Date"].between(cfg.START_DATE, cfg.END_DATE)].reset_index(drop=True)
    df = df.dropna(subset=["Price"]).reset_index(drop=True)
    return df


def _add_time_index(df: pd.DataFrame) -> pd.DataFrame:
    """Add integer time index ``t`` and calendar month key."""
    df["t"] = range(len(df))
    _dt = pd.to_datetime(df["Date"])
    df["month_key"] = _dt.dt.to_period("M").astype(str)
    df["year"] = _dt.dt.year
    df["month_num"] = _dt.dt.month
    return df


def _find_col(df: pd.DataFrame, pattern: str, block: str) -> pd.Series:
    """Find a block-suffixed column by case-insensitive pattern match."""
    pat_low = pattern.lower()
    candidates = [
        c for c in df.columns
        if pat_low in c.lower() and c.endswith(f"_{block}")
    ]
    if candidates:
        return pd.to_numeric(df[candidates[0]], errors="coerce").fillna(0.0)
    return pd.Series(0.0, index=df.index, dtype=float)


def _add_unavailability_block(df: pd.DataFrame, block: str) -> pd.DataFrame:
    """Detect and normalise the unavailability column for one block."""
    col = _find_col(df, "unavail", block)
    df[f"unavailibility_{block}"] = col.fillna(0).round().astype(int)
    return df


def _add_pmin_pmax_block(df: pd.DataFrame, block: str) -> pd.DataFrame:
    """Read per-block Pmin / Pmax time series."""
    # Look for columns like "Pmin_A", "Pmax_A" (already suffixed from _read_input)
    pmin_col = _find_col(df, "pmin", block)
    pmax_col = _find_col(df, "pmax", block)
    df[f"Pmin_{block}"] = pmin_col
    df[f"Pmax_{block}"] = pmax_col
    return df


def _add_gridfee(df: pd.DataFrame) -> pd.DataFrame:
    """Detect and normalise the grid-fee time series."""
    gridfee_cols = [
        c
        for c in df.columns
        if (
            ("grid" in c.lower() and "fee" in c.lower())
            or c.strip().lower()
            in [
                "gridfee", "grid_fee", "grid fee",
                "grid_fee_eur_mwh", "grid fee eur/mwh", "grid fee (eur/mwh)",
            ]
        )
    ]
    if gridfee_cols:
        df["GRIDFEE"] = (
            pd.to_numeric(df[gridfee_cols[0]], errors="coerce")
            .fillna(cfg.DEFAULT_GRIDFEE)
        )
    else:
        df["GRIDFEE"] = cfg.DEFAULT_GRIDFEE
    return df


def _add_season_and_dow(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare DOW column and revenues."""
    # Rename input column Warme → DOW if present
    if "Warme" in df.columns:
        df = df.rename(columns={"Warme": "DOW"})
    df["DOW revenues"] = 0.0

    return df


def _compute_cost_curves(df: pd.DataFrame) -> Tuple[pd.DataFrame, dict]:
    """
    Build per-block linear total-cost approximation between Pmin(t) and Pmax(t).

    For each block b ∈ {A, B}:
      total_cost_b(t) = cost_slope_b(t) * P_eff_b(t) + cost_fixed_b(t) * on_b(t)

    Returns the enriched DataFrame *and* a dict with per-block Pmax_eff values.
    """
    cost_meta: dict = {}

    for b in cfg.BLOCKS:
        # Ensure Pmin / Pmax and unavailability columns exist
        df = _add_pmin_pmax_block(df, b)
        df = _add_unavailability_block(df, b)

    # ---- Cost curves use ORIGINAL Pmax (before DOW reduction) ----
    # The plant burns fuel for the full thermal load regardless of DOW;
    # DOW only limits how much electricity can be sold.
    for b in cfg.BLOCKS:
        pmin = df[f"Pmin_{b}"]
        pmax = df[f"Pmax_{b}"]  # original from input

        # TC at Pmin / Pmax (look for block-suffixed columns)
        tc_pmin = _find_col(df, "total generation costs at pmin", b)
        tc_pmax = _find_col(df, "total generation costs at pmax", b)
        df[f"TC_PminN_{b}"] = tc_pmin
        df[f"TC_Pmax_{b}"] = tc_pmax

        # Pmax_eff = original Pmax (P_eff can reach P + DOW = original Pmax)
        cost_meta[f"Pmax_eff_{b}"] = float(pmax.max())

        # Single linear segment per hour: cost = slope * P_eff + fixed * on
        Cmin = tc_pmin * pmin
        Cmax = tc_pmax * pmax
        denom = (pmax - pmin).replace(0.0, np.nan)
        df[f"cost_slope_{b}"] = (Cmax - Cmin) / denom
        df[f"cost_slope_{b}"] = df[f"cost_slope_{b}"].fillna(0.0)
        df[f"cost_fixed_{b}"] = Cmin - df[f"cost_slope_{b}"] * pmin

    # DOW Pmax reduction is now handled dynamically in model constraints
    # (depends on which block is actually dispatched, not just availability).
    # Pmax stays at original values here; model_builder applies DOW deduction.

    return df, cost_meta


def _compute_coal_curves(df: pd.DataFrame) -> pd.DataFrame:
    """Compute per-block coal consumption linear coefficients.

    coal_consumption(P_eff) = coal_slope * P_eff + coal_fixed * on  [t/h]

    Uses the same Pmin/Pmax interpolation as the cost curves.
    """
    for b in cfg.BLOCKS:
        pmin = df[f"Pmin_{b}"]
        pmax = df[f"Pmax_{b}"]

        coal_pmin = _find_col(df, "coal conversion factor at pmin", b)  # t/MWh
        coal_pmax = _find_col(df, "coal conversion factor at pmax", b)  # t/MWh

        # Total coal at Pmin / Pmax [t/h]
        C_coal_min = coal_pmin * pmin
        C_coal_max = coal_pmax * pmax

        denom = (pmax - pmin).replace(0.0, np.nan)
        df[f"coal_slope_{b}"] = ((C_coal_max - C_coal_min) / denom).fillna(0.0)
        df[f"coal_fixed_{b}"] = C_coal_min - df[f"coal_slope_{b}"] * pmin

    return df


# ====================================================================
#  Coal constraints tab – monthly coal volume limits
# ====================================================================


def _read_coal_constrains_tab() -> dict:
    """Read monthly coal volume limits from the *Coal_constrains* tab.

    Returns
    -------
    dict
        ``{(year, month): limit_kt}`` where limit_kt is in kilotonnes.
        Empty dict if the tab does not exist.
    """
    import shutil
    import tempfile

    from openpyxl import load_workbook

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    shutil.copy2(cfg.INPUT_FILE, tmp_path)

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        if "Coal_constrains" not in wb.sheetnames:
            wb.close()
            print("--- Coal_constrains tab not found; skipping coal limits")
            return {}
        ws = wb["Coal_constrains"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        wb.close()
    finally:
        os.unlink(tmp_path)

    limits: dict = {}
    for row in rows:
        if row[0] is None or row[1] is None:
            continue
        year = int(row[0])
        raw_month = row[1]
        try:
            month = int(raw_month)
        except (ValueError, TypeError):
            # Month given as name/abbreviation (e.g. 'Mar' or 'March')
            try:
                month = pd.to_datetime(str(raw_month), format="%b").month
            except ValueError:
                month = pd.to_datetime(str(raw_month), format="%B").month
        limit_kt = float(row[2]) if row[2] is not None else None
        if limit_kt is not None:
            limits[(year, month)] = limit_kt

    if limits:
        print(f"--- Coal constraints loaded: {len(limits)} months")
        for (y, m_), lim in sorted(limits.items()):
            print(f"    {y}-{m_:02d}: {lim:.1f} kt")

    return limits


# ====================================================================
#  Starts tab – tiered startup costs & ramp profiles
# ====================================================================


def _read_starts_tab() -> dict:
    """Read per-block startup tiers (costs & ramp MW) from the *Starts* tab.

    Returns
    -------
    dict
        ``{block: [tier_dict, ...]}`` where each tier dict has keys:
        ``name``, ``min_off``, ``max_off``, ``cost``, ``ramp`` (list of MW).
    """
    import os
    import shutil
    import tempfile

    from openpyxl import load_workbook

    # Copy to a temp file to avoid Excel lock
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    shutil.copy2(cfg.INPUT_FILE, tmp_path)

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb["Starts"]
        rows = list(ws.iter_rows(min_row=1, max_row=20, max_col=6, values_only=True))
        wb.close()
    finally:
        os.unlink(tmp_path)

    tier_names = ["very_hot", "hot", "warm", "cold", "vcold"]
    tier_bounds = [(0, 5), (5, 10), (10, 60), (60, 100), (100, None)]

    # Block A: data rows 3-7 (0-indexed 2-6), Block B: rows 11-15 (0-indexed 10-14)
    block_offsets = {"A": 2, "B": 10}

    starts: dict = {}
    for block in cfg.BLOCKS:
        offset = block_offsets[block]
        tiers = []
        for i, (tname, (lo, hi)) in enumerate(zip(tier_names, tier_bounds)):
            row = rows[offset + i]
            ramp = [float(v) for v in row[1:5] if v is not None]
            cost = float(row[5]) if row[5] is not None else 0.0
            tiers.append({
                "name": tname,
                "min_off": lo,
                "max_off": hi,
                "cost": cost,
                "ramp": ramp,
            })
        starts[block] = tiers

    return starts
