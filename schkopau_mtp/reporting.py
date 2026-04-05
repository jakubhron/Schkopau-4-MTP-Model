"""
Excel reporting for the Schkopau MTP model.

Writes the two-sheet workbook:
  - **Results** : full hourly time-series with audit columns
  - **Monthly** : styled monthly P&L / operations summary
"""

from __future__ import annotations

import numbers
import re
from typing import List, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config as cfg


# ====================================================================
#  PUBLIC API
# ====================================================================


def write_excel(
    df: pd.DataFrame,
    cost_meta: dict,
    output_path: str,
    *,
    coal_shadow_prices: dict | None = None,
) -> None:
    """
    Create the output Excel workbook with a *Results* and *Monthly* sheet.

    Parameters
    ----------
    df : pd.DataFrame  – hourly data with PnL columns.
    cost_meta : dict    – contains Pmax_eff.
    output_path : str   – destination xlsx path.
    coal_shadow_prices : dict, optional
        {(year, month): EUR/t} shadow prices from the coal constraint LP re-solve.
    """
    Pmax_eff = max(cost_meta.get(f"Pmax_eff_{b}", 0.0) for b in cfg.BLOCKS)

    # We modify a working copy so that the caller's df is not mutated.
    df_m = df.copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # --- Sheet 1: full hourly results (initial) ---
        df.to_excel(writer, index=False, sheet_name="Results")

        # --- Enrich df_m with reporting columns ---
        df_m = _prepare_monthly_columns(df_m, Pmax_eff)

        # Write audit columns back to Results
        audit_cols = [
            "CO2_factor_tCO2_per_MWh", "Coal_factor_t_per_MWh",
            "Coal_t_h", "CO2_t_h",
            "CO2_cost_EUR_h",
            "Merchant_power_revenue_EUR_h",
            "DOW_AP_revenue_EUR_h", "DOW_CHP_subsidy_EUR_h",
            "Coal_API2_cost_EUR_h", "Coal_other_cost_EUR_h",
            "Start_cost_EUR_h", "Variable_other_cost_EUR_h",
            "TC_stack_delta_EUR_h",
            "House_power_EUR_h",
        ]
        present = [c for c in audit_cols if c in df_m.columns]
        df_out = pd.concat([df, df_m[present].copy()], axis=1)

        # Write hourly cost / MW splits to Results as well
        monthly_cols = [
            "Merchant_GWh_h",
            "Coal_kMT_h_Merchant", "Coal_kMT_h_DOW",
            "CO2_kMT_h_Merchant", "CO2_kMT_h_DOW",
            "Coal_API2_EUR_h_Merchant", "Coal_API2_EUR_h_DOW",
            "Coal_other_EUR_h_Merchant", "Coal_other_EUR_h_DOW",
            "CO2_EUR_h_Merchant", "CO2_EUR_h_DOW",
            "Coal_kMT_h", "CO2_kMT_h",
            "DOW_GWhth_h",
        ]
        present = [c for c in monthly_cols if c in df_m.columns]
        df_out = pd.concat([df_out, df_m[present]], axis=1)

        # Overwrite Results
        df_out.to_excel(writer, index=False, sheet_name="Results")

        # --- Sheet 2: Monthly summary (plant total) ---
        _write_monthly_sheet(writer, df_m, output_path, sheet_name="Monthly",
                             coal_shadow_prices=coal_shadow_prices)

        # --- Per-block Monthly sheets ---
        for blk in cfg.BLOCKS:
            df_blk = df.copy()
            Pmax_eff_blk = cost_meta.get(f"Pmax_eff_{blk}", 0.0)
            df_blk = _prepare_monthly_columns(df_blk, Pmax_eff_blk, block=blk)
            _write_monthly_sheet(
                writer, df_blk, output_path,
                sheet_name=f"Monthly_{blk}", block=blk,
                coal_shadow_prices=coal_shadow_prices,
            )

    print(f"\n✓ {cfg.VERSION} saved as:")
    print(output_path)


# ====================================================================
#  INTERNAL – emission / cost factor helpers
# ====================================================================


def _linear_cost(x, y1, y2, p1, p2):
    """Linear interpolation of TOTAL COST between (p1, y1) and (p2, y2).

    Values at or below p1 get y1; values at or above p2 get y2.
    """
    x = pd.to_numeric(x, errors="coerce").fillna(0.0).clip(lower=0.0)
    denom = p2 - p1
    if isinstance(denom, pd.Series):
        denom = denom.replace(0.0, np.nan)
    elif denom == 0:
        denom = np.nan
    x_clamped = x.clip(lower=p1, upper=p2) if isinstance(p1, pd.Series) else x.clip(lower=p1, upper=p2)
    result = y1 + (y2 - y1) * (x_clamped - p1) / denom
    return result.fillna(0.0)


def _s(df_m, col, default=0.0):
    """Return numeric Series for *col* if present, else a Series of *default*."""
    if col in df_m.columns:
        return pd.to_numeric(df_m[col], errors="coerce").fillna(default)
    return pd.Series(default, index=df_m.index, dtype=float)


# ====================================================================
#  INTERNAL – enrich DataFrame with monthly-reporting columns
# ====================================================================


def _prepare_monthly_columns(
    df_m: pd.DataFrame,
    Pmax_eff: float,
    block: str | None = None,
) -> pd.DataFrame:
    """Add all columns needed for the Monthly sheet.

    Parameters
    ----------
    block : str or None
        If ``"A"`` or ``"B"``, compute columns for that single block.
        If *None*, compute plant-level totals (default / existing behaviour).
    """
    # Block suffix for input/factor column lookups
    bsf = f"_{block}" if block else "_A"

    # --- When per-block, remap plant-level columns to block-specific ---
    if block:
        df_m["P"] = _s(df_m, f"P_{block}", 0.0)
        df_m["on_model"] = _s(df_m, f"on_model_{block}", 0.0).round().astype(int)
        df_m["start_cost"] = _s(df_m, f"start_cost_{block}", 0.0)
        # OFF_costs is plant-level (both-off only); split evenly for per-block view
        df_m["OFF_costs"] = _s(df_m, "OFF_costs", 0.0) / len(cfg.BLOCKS)
        # DOW attribution: dispatch-based (primary serves when on, else other)
        _on_primary = _s(df_m, f"on_model_{cfg.DOW_BLOCK}", 0.0).round().astype(int)
        if block == cfg.DOW_BLOCK:
            _dow_frac = _on_primary
        else:
            _on_this = _s(df_m, f"on_model_{block}", 0.0).round().astype(int)
            _dow_frac = (_on_this * (1 - _on_primary))
        # Scale plant-level DOW to this block's share
        _plant_dow_on = (_on_primary + _s(df_m, f"on_model_{[b for b in cfg.BLOCKS if b != cfg.DOW_BLOCK][0]}", 0.0).round().astype(int)).clip(upper=1)
        _frac_safe = _dow_frac / _plant_dow_on.replace(0, 1)  # avoid /0 when both off
        df_m["DOW_revenues_real"] = _s(df_m, "DOW_revenues_real", 0.0) * _frac_safe
        df_m["dow_subsidy"] = _s(df_m, "dow_subsidy", 0.0) * _frac_safe
        df_m["DOW revenues"] = _s(df_m, "DOW revenues", 0.0) * _frac_safe
        df_m["PnL"] = (
            _s(df_m, f"profit_spot_{block}", 0.0)
            + _s(df_m, "DOW_revenues_real", 0.0)
            + _s(df_m, "dow_subsidy", 0.0)
            - _s(df_m, f"run_costs_{block}", 0.0)
            - _s(df_m, f"start_cost_{block}", 0.0)
            - df_m["OFF_costs"]
        )

    # Per-block Pmin/Pmax time series
    P1 = _s(df_m, f"Pmin{bsf}", 0.0)
    P2 = _s(df_m, f"Pmax{bsf}", 0.0)

    # --- ON status ---
    df_m["Date"] = pd.to_datetime(df_m["Date"])
    df_m["Year"] = df_m["Date"].dt.year
    df_m["MonthNum"] = df_m["Date"].dt.month

    on_src = f"on_model_{block}" if block else "on_model"
    df_m["on"] = (
        pd.to_numeric(df_m.get(on_src, df_m.get("on", 0)), errors="coerce")
        .fillna(0.0)
        .round()
        .astype(int)
    )
    _unavail_report = (
        pd.to_numeric(df_m.get(f"unavailibility{bsf}", df_m.get("unavailibility", 0)), errors="coerce")
        .fillna(0)
        .round()
        .astype(int)
    )
    df_m["on"] = (df_m["on"] * (1 - _unavail_report)).astype(int)

    P_val = pd.to_numeric(df_m.get("P", 0.0), errors="coerce").fillna(0.0)
    df_m["P_eff"] = P_val

    P_eff_num = pd.to_numeric(df_m.get("P_eff", 0.0), errors="coerce").fillna(0.0)

    # --- Helper: per-block emission / cost stack computation ---
    def _compute_block_costs(df_m, blk_suffix, p_eff_blk, p1_blk, p2_blk):
        """Compute cost stack for one block; return dict of Series."""
        # Prefer *actual* emission-factor columns (tCO2/MWh ~1.0) over
        # "CO2-Emmission factor" columns which in some inputs contain
        # pre-computed EUR/MWh costs (factor × EUA).
        _co2_pminn = pd.to_numeric(
            df_m.get(f"Emission factor at Pmin [tCO2/MWh]{blk_suffix}",
                df_m.get(f"CO2-Emmission factor at Pmin_N [tCO2/MWh]{blk_suffix}",
                    df_m.get(f"CO2-Emmission factor at Pmin [tCO2/MWh]{blk_suffix}",
                        df_m.get("Emission factor at Pmin [tCO2/MWh]",
                            df_m.get("CO2-Emmission factor at Pmin_N [tCO2/MWh]",
                                df_m.get("CO2-Emmission factor at Pmin [tCO2/MWh]", 0.0)))))),
            errors="coerce",
        ).fillna(0.0)
        _co2_pmax = pd.to_numeric(
            df_m.get(f"Emission factor at Pmax [tCO2/GJ]{blk_suffix}",
                df_m.get(f"Emission factor at Pmax [tCO2/MWh]{blk_suffix}",
                    df_m.get(f"CO2-Emmission factor at Pmax [tCO2/MWh]{blk_suffix}",
                        df_m.get("Emission factor at Pmax [tCO2/GJ]",
                            df_m.get("Emission factor at Pmax [tCO2/MWh]",
                                df_m.get("CO2-Emmission factor at Pmax [tCO2/MWh]", 0.0)))))),
            errors="coerce",
        ).fillna(0.0)
        _coal_pminn = pd.to_numeric(
            df_m.get(f"Coal conversion factor at Pmin_N [t/MWh]{blk_suffix}",
                      df_m.get(f"Coal conversion factor at Pmin [t/MWh]{blk_suffix}",
                               df_m.get("Coal conversion factor at Pmin_N [t/MWh]",
                                        df_m.get("Coal conversion factor at Pmin [t/MWh]", 0.0)))),
            errors="coerce",
        ).fillna(0.0)
        _coal_pmax = pd.to_numeric(
            df_m.get(f"Coal conversion factor at Pmax [t/MWh]{blk_suffix}",
                     df_m.get("Coal conversion factor at Pmax [t/MWh]", 0.0)),
            errors="coerce",
        ).fillna(0.0)

        def _pw2_blk(x, y1, y2, pp1, pp2):
            x = pd.to_numeric(x, errors="coerce").fillna(0.0).clip(lower=pp1, upper=pp2)
            denom = (pp2 - pp1).replace(0.0, np.nan) if isinstance(pp2, pd.Series) else (pp2 - pp1 if pp2 != pp1 else np.nan)
            return (y1 + (y2 - y1) * (x - pp1) / denom).fillna(0.0)

        co2_factor = _pw2_blk(p_eff_blk, _co2_pminn, _co2_pmax, p1_blk, p2_blk)
        coal_factor = _pw2_blk(p_eff_blk, _coal_pminn, _coal_pmax, p1_blk, p2_blk)

        _api2_col = (
            "API2 EUR/t" if "API2 EUR/t" in df_m.columns
            else ("API2" if "API2" in df_m.columns
                  else ("Coal Price" if "Coal Price" in df_m.columns else None))
        )
        _api2_p = pd.to_numeric(df_m[_api2_col], errors="coerce").fillna(0.0) if _api2_col else 0.0
        _tr_col = next((c for c in df_m.columns if "coal transportation" in c.lower()), None)
        _tr_p = pd.to_numeric(df_m[_tr_col], errors="coerce").fillna(0.0) if _tr_col else 0.0
        _eua_p = pd.to_numeric(df_m.get("EUA", 0.0), errors="coerce").fillna(0.0)

        y1_api2 = _coal_pminn * _api2_p * p1_blk
        y2_api2 = _coal_pmax * _api2_p * p2_blk
        coal_api2 = _linear_cost(p_eff_blk, y1_api2, y2_api2, p1_blk, p2_blk)

        y1_tr = _coal_pminn * _tr_p * p1_blk
        y2_tr = _coal_pmax * _tr_p * p2_blk
        coal_other = _linear_cost(p_eff_blk, y1_tr, y2_tr, p1_blk, p2_blk)

        y1_co2 = _co2_pminn * _eua_p * p1_blk
        y2_co2 = _co2_pmax * _eua_p * p2_blk
        co2_cost = _linear_cost(p_eff_blk, y1_co2, y2_co2, p1_blk, p2_blk)

        _oth_col = next((c for c in df_m.columns if "total additional costs" in c.lower()), None)
        _oth_mwh = pd.to_numeric(df_m[_oth_col], errors="coerce").fillna(0.0) if _oth_col else 0.0
        var_other = _oth_mwh * p_eff_blk

        return {
            "co2_factor": co2_factor, "coal_factor": coal_factor,
            "coal_t": coal_factor * p_eff_blk, "co2_t": co2_factor * p_eff_blk,
            "coal_api2": coal_api2, "coal_other": coal_other,
            "co2_cost": co2_cost, "var_other": var_other,
            "api2_price": _api2_p, "transport_price": _tr_p,
        }

    # --- Compute cost stacks per block and aggregate ---
    blocks_to_compute = [block] if block else cfg.BLOCKS

    # Accumulators
    _coal_t_acc = pd.Series(0.0, index=df_m.index)
    _co2_t_acc = pd.Series(0.0, index=df_m.index)
    _coal_api2_acc = pd.Series(0.0, index=df_m.index)
    _coal_other_acc = pd.Series(0.0, index=df_m.index)
    _co2_cost_acc = pd.Series(0.0, index=df_m.index)
    _var_other_acc = pd.Series(0.0, index=df_m.index)
    _rc_acc = pd.Series(0.0, index=df_m.index)
    _last_api2_price = 0.0
    _last_transport_price = 0.0

    for _b in blocks_to_compute:
        _bsf_b = f"_{_b}"
        _p1_b = _s(df_m, f"Pmin{_bsf_b}", 0.0)
        _p2_b = _s(df_m, f"Pmax{_bsf_b}", 0.0)
        _peff_b = _s(df_m, f"P_eff_{_b}", 0.0)
        _on_b = pd.to_numeric(df_m.get(f"on_model_{_b}", 0), errors="coerce").fillna(0.0)

        bc = _compute_block_costs(df_m, _bsf_b, _peff_b, _p1_b, _p2_b)
        _coal_t_acc += bc["coal_t"]
        _co2_t_acc += bc["co2_t"]
        _coal_api2_acc += bc["coal_api2"]
        _coal_other_acc += bc["coal_other"]
        _co2_cost_acc += bc["co2_cost"]
        _var_other_acc += bc["var_other"]
        _last_api2_price = bc["api2_price"]
        _last_transport_price = bc["transport_price"]

        _cs = pd.to_numeric(df_m.get(f"cost_slope_{_b}", 0.0), errors="coerce").fillna(0.0)
        _cf = pd.to_numeric(df_m.get(f"cost_fixed_{_b}", 0.0), errors="coerce").fillna(0.0)
        _rc_acc += _cs * _peff_b + _cf * _on_b

    # Use single-block factor columns for per-block view (reporting display)
    if block:
        _bsf_disp = f"_{block}"
        _peff_disp = _s(df_m, f"P_eff_{block}", P_eff_num)
        _p1_disp = _s(df_m, f"Pmin{_bsf_disp}", 0.0)
        _p2_disp = _s(df_m, f"Pmax{_bsf_disp}", 0.0)
        bc_disp = _compute_block_costs(df_m, _bsf_disp, _peff_disp, _p1_disp, _p2_disp)
        df_m["CO2_factor_tCO2_per_MWh"] = bc_disp["co2_factor"]
        df_m["Coal_factor_t_per_MWh"] = bc_disp["coal_factor"]
    else:
        # Plant average factors (weighted)
        _total_peff = sum(_s(df_m, f"P_eff_{_b}", 0.0) for _b in cfg.BLOCKS)
        _total_peff_safe = _total_peff.replace(0.0, np.nan)
        df_m["CO2_factor_tCO2_per_MWh"] = (_co2_t_acc / _total_peff_safe).fillna(0.0)
        df_m["Coal_factor_t_per_MWh"] = (_coal_t_acc / _total_peff_safe).fillna(0.0)

    df_m["RunCosts_TC_EUR_h"] = _rc_acc
    df_m["Coal_t_h"] = _coal_t_acc
    df_m["CO2_t_h"] = _co2_t_acc

    # Hourly revenues / costs
    df_m["Merchant_power_revenue_EUR_h"] = df_m["P"] * df_m["Price"]
    df_m["DOW_AP_revenue_EUR_h"] = df_m.get("DOW_revenues_real", 0.0)
    df_m["DOW_CHP_subsidy_EUR_h"] = df_m.get("dow_subsidy", 0.0)

    df_m["Coal_API2_cost_EUR_h"] = _coal_api2_acc
    df_m["Coal_other_cost_EUR_h"] = _coal_other_acc
    df_m["CO2_cost_EUR_h"] = _co2_cost_acc
    df_m["Variable_other_cost_EUR_h"] = _var_other_acc

    api2_price = _last_api2_price
    transport_price = _last_transport_price

    df_m["TC_stack_delta_EUR_h"] = (
        _rc_acc
        - _coal_api2_acc
        - _coal_other_acc
        - _co2_cost_acc
        - _var_other_acc
    )

    # --- Implied cost split from solver running costs ---
    stack_implied = _coal_api2_acc + _coal_other_acc + _co2_cost_acc + _var_other_acc
    _run_tc = _rc_acc

    def _share(component):
        return np.where(stack_implied > 0.0, component / stack_implied, 0.0)

    sh_coal_api2 = _share(_coal_api2_acc)
    sh_coal_other = _share(_coal_other_acc)
    sh_co2 = _share(_co2_cost_acc)
    sh_other = _share(_var_other_acc)

    df_m["Coal_API2_cost_solver_EUR_h"] = sh_coal_api2 * _run_tc
    df_m["Coal_other_cost_solver_EUR_h"] = sh_coal_other * _run_tc
    df_m["CO2_cost_solver_EUR_h"] = sh_co2 * _run_tc
    df_m["Variable_other_cost_solver_EUR_h"] = sh_other * _run_tc

    df_m["TC_stack_delta_EUR_h"] = _run_tc - (
        df_m["Coal_API2_cost_solver_EUR_h"]
        + df_m["Coal_other_cost_solver_EUR_h"]
        + df_m["CO2_cost_solver_EUR_h"]
        + df_m["Variable_other_cost_solver_EUR_h"]
    )

    # Implied quantities
    _eua = pd.to_numeric(df_m.get("EUA", 0.0), errors="coerce").fillna(0.0)
    df_m = df_m.copy()
    df_m["CO2_t_h_implied_from_solver"] = np.where(
        _eua > 0.0, df_m["CO2_cost_solver_EUR_h"] / _eua, 0.0
    )
    coal_cost_solver_total = (
        df_m["Coal_API2_cost_solver_EUR_h"] + df_m["Coal_other_cost_solver_EUR_h"]
    )
    coal_price_total = (
        pd.to_numeric(api2_price, errors="coerce").fillna(0.0)
        + pd.to_numeric(transport_price, errors="coerce").fillna(0.0)
    )
    df_m["Coal_t_h_implied_from_solver"] = np.where(
        coal_price_total > 0.0, coal_cost_solver_total / coal_price_total, 0.0
    )

    df_m["Start_cost_EUR_h"] = df_m.get("start_cost", 0.0)
    df_m["House_power_EUR_h"] = df_m.get("OFF_costs", 0.0)

    df_m["Merchant_GWh_h"] = df_m["P"] / 1000.0

    # DOW GWh/h (no electrical equivalent — DOW reduces Pmax directly)
    _on = _s(df_m, "on_model", _s(df_m, "on", 0.0))
    _dow_MW = _s(df_m, "DOW", 0.0).clip(lower=0.0) * (_on >= 0.5).astype(float)
    df_m["DOW_GWh_h"] = _dow_MW / 1000.0

    # --- Merchant / DOW split (no must-run / profitable breakdown) ---
    _price = _s(df_m, "Price", 0.0)
    _merchant_MW = _s(df_m, "P", 0.0).clip(lower=0.0)

    df_m["Merchant_GWh_h"] = _merchant_MW / 1000.0

    _den = (_merchant_MW + _dow_MW).replace(0.0, np.nan)

    _coal_t_h = _s(df_m, "Coal_t_h_implied_from_solver", _s(df_m, "Coal_t_h", 0.0))
    _co2_t_h = _s(df_m, "CO2_t_h_implied_from_solver", _s(df_m, "CO2_t_h", 0.0))

    df_m["Coal_kMT_h_Merchant"] = (_coal_t_h * (_merchant_MW / _den)).fillna(0.0) / 1000.0
    df_m["Coal_kMT_h_DOW"] = (_coal_t_h * (_dow_MW / _den)).fillna(0.0) / 1000.0

    df_m["CO2_kMT_h_Merchant"] = (_co2_t_h * (_merchant_MW / _den)).fillna(0.0) / 1000.0
    df_m["CO2_kMT_h_DOW"] = (_co2_t_h * (_dow_MW / _den)).fillna(0.0) / 1000.0

    # EUR breakdown
    _cost_api2 = _s(df_m, "Coal_API2_cost_solver_EUR_h", _s(df_m, "Coal_API2_cost_EUR_h", 0.0))
    _cost_other = _s(df_m, "Coal_other_cost_solver_EUR_h", _s(df_m, "Coal_other_cost_EUR_h", 0.0))
    _cost_co2 = _s(df_m, "CO2_cost_solver_EUR_h", _s(df_m, "CO2_cost_EUR_h", 0.0))

    df_m["Coal_API2_EUR_h_Merchant"] = (_cost_api2 * (_merchant_MW / _den)).fillna(0.0)
    df_m["Coal_API2_EUR_h_DOW"] = (_cost_api2 * (_dow_MW / _den)).fillna(0.0)

    df_m["Coal_other_EUR_h_Merchant"] = (_cost_other * (_merchant_MW / _den)).fillna(0.0)
    df_m["Coal_other_EUR_h_DOW"] = (_cost_other * (_dow_MW / _den)).fillna(0.0)

    df_m["CO2_EUR_h_Merchant"] = (_cost_co2 * (_merchant_MW / _den)).fillna(0.0)
    df_m["CO2_EUR_h_DOW"] = (_cost_co2 * (_dow_MW / _den)).fillna(0.0)

    df_m["Coal_kMT_h"] = _coal_t_h.fillna(0.0) / 1000.0
    df_m["CO2_kMT_h"] = _co2_t_h.fillna(0.0) / 1000.0
    df_m = df_m.copy()

    onm = _on
    u = _s(df_m, f"unavailibility{bsf}", _s(df_m, "unavailibility", 0.0))
    w = _s(df_m, "DOW", 0.0)
    df_m["DOW_GWhth_h"] = (w * onm * (1 - u)) / 1000.0

    return df_m


# ====================================================================
#  INTERNAL – Monthly summary sheet
# ====================================================================


def _write_monthly_sheet(
    writer,
    df_m: pd.DataFrame,
    output_path: str,
    sheet_name: str = "Monthly",
    block: str | None = None,
    coal_shadow_prices: dict | None = None,
) -> None:
    """Build the Monthly aggregation table and style it."""
    bsf = f"_{block}" if block else "_A"

    def msum(series_name):
        if series_name in df_m.columns:
            return df_m.groupby(["Year", "MonthNum"])[series_name].sum()
        return pd.Series(dtype=float)

    def mmean(series_name):
        if series_name in df_m.columns:
            return df_m.groupby(["Year", "MonthNum"])[series_name].mean()
        return pd.Series(dtype=float)

    periods = list(
        df_m[["Year", "MonthNum"]]
        .drop_duplicates()
        .sort_values(["Year", "MonthNum"])
        .itertuples(index=False, name=None)
    )

    mv: dict = {}  # monthly_values

    # ---- Revenues ----
    mv["EPEX Revenues"] = msum("Merchant_power_revenue_EUR_h")
    mv["DOW AP revenue"] = msum("DOW_AP_revenue_EUR_h")
    mv["DOW CHP subsidy"] = msum("DOW_CHP_subsidy_EUR_h")

    mv["DOW revenues (TOTAL)"] = mv["DOW AP revenue"].add(
        mv["DOW CHP subsidy"], fill_value=0.0
    )
    mv["Total Revenues (TOTAL)"] = (
        mv["EPEX Revenues"]
        .add(mv["DOW revenues (TOTAL)"], fill_value=0.0)
    )

    # ---- Costs ----
    mv["Coal API2"] = (
        msum("Coal_API2_cost_solver_EUR_h")
        if "Coal_API2_cost_solver_EUR_h" in df_m.columns
        else msum("Coal_API2_cost_EUR_h")
    )
    mv["Coal other costs (premium, transportation)"] = (
        msum("Coal_other_cost_solver_EUR_h")
        if "Coal_other_cost_solver_EUR_h" in df_m.columns
        else msum("Coal_other_cost_EUR_h")
    )
    mv["CO2 costs"] = (
        msum("CO2_cost_solver_EUR_h")
        if "CO2_cost_solver_EUR_h" in df_m.columns
        else msum("CO2_cost_EUR_h")
    )

    # Cost breakdowns: Merchant + DOW only
    for prefix, base_col in [
        ("Coal API2", "Coal_API2_EUR_h"),
        ("Coal other", "Coal_other_EUR_h"),
        ("CO2 costs", "CO2_EUR_h"),
    ]:
        for suffix in ["Merchant", "DOW"]:
            key = f"{prefix} -- {suffix}"
            col_name = f"{base_col}_{suffix}"
            mv[key] = msum(col_name)

    mv["Start costs"] = msum("Start_cost_EUR_h")
    mv["Variable other costs"] = (
        msum("Variable_other_cost_solver_EUR_h")
        if "Variable_other_cost_solver_EUR_h" in df_m.columns
        else msum("Variable_other_cost_EUR_h")
    )
    mv["TC stack delta costs"] = msum("TC_stack_delta_EUR_h")
    mv["House Power"] = msum("House_power_EUR_h")

    mv["Other costs (TOTAL)"] = (
        mv["Start costs"]
        .add(mv["Variable other costs"], fill_value=0.0)
        .add(mv["House Power"], fill_value=0.0)
    )
    mv["Total Costs (TOTAL)"] = (
        mv["Coal API2"]
        .add(mv["Coal other costs (premium, transportation)"], fill_value=0.0)
        .add(mv["CO2 costs"], fill_value=0.0)
        .add(mv["TC stack delta costs"], fill_value=0.0)
        .add(mv["Other costs (TOTAL)"], fill_value=0.0)
    )
    mv["PnL"] = mv["Total Revenues (TOTAL)"].sub(mv["Total Costs (TOTAL)"], fill_value=0.0)

    # Solver PnL total
    mv["Total"] = msum("PnL")

    # Revenue breakdown (no must-run/profitable split)

    # ---- Operation statistics ----
    mv["Power produced Merchant"] = msum("Merchant_GWh_h")
    mv["Power produced DOW"] = msum("DOW_GWh_h")
    mv["Coal Consumption"] = msum("Coal_kMT_h")
    mv["CO2 emissions"] = msum("CO2_kMT_h")

    for resource, prefix_col in [("Coal Consumption", "Coal_kMT_h"), ("CO2 emissions", "CO2_kMT_h")]:
        for suffix in ["Merchant", "DOW"]:
            mv[f"{resource} -- {suffix}"] = msum(f"{prefix_col}_{suffix}")

    # Counts
    if "start_cost" in df_m.columns:
        start_mask = pd.to_numeric(df_m["start_cost"], errors="coerce").fillna(0.0) > 0
    else:
        start_mask = pd.Series(False, index=df_m.index)
    outage_mask = pd.to_numeric(
        df_m.get(f"unavailibility{bsf}", df_m.get("unavailibility", 0.0)), errors="coerce"
    ).fillna(0.0) == 1

    mv["Number of starts"] = (
        start_mask.astype(int)
        .groupby([df_m["Year"], df_m["MonthNum"]])
        .sum()
        .astype(int)
    )
    mv["Outage"] = (
        outage_mask.astype(int)
        .groupby([df_m["Year"], df_m["MonthNum"]])
        .sum()
        .astype(int)
    )
    mv["DOW volume"] = msum("DOW_GWhth_h").round().astype(int)

    # ---- Coal shadow prices (from LP re-solve) ----
    if coal_shadow_prices:
        mv["Coal shadow price"] = pd.Series(
            {(y, mo): v for (y, mo), v in coal_shadow_prices.items()},
            dtype=float,
        )

    # ---- Price averages ----
    mv["EPEX forecast BL"] = mmean("Price")
    mv["CDS"] = mmean("Price").sub(mmean(f"TC_Pmax{bsf}"), fill_value=0.0)

    # ----------------------------------------------------------------
    #  Build the output table
    # ----------------------------------------------------------------
    output_rows = _define_output_rows()

    data = []
    for label, unit, key in output_rows:
        row: list = [label, unit]
        for yy, mm in periods:
            if key is None:
                row.append("")
            else:
                s = mv.get(key)
                val = 0.0
                if s is not None and (yy, mm) in s.index:
                    val = float(s.loc[(yy, mm)])
                if unit in ("#", "h", "GWhth"):
                    val = int(round(val))
                row.append(val)
        data.append(row)

    monthly_df = pd.DataFrame(
        data,
        columns=["Item", "Unit"] + [f"{yy}-{mm:02d}" for yy, mm in periods],
    )
    monthly_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)

    # ----------------------------------------------------------------
    #  Style the Monthly sheet
    # ----------------------------------------------------------------
    ws = writer.sheets[sheet_name]
    _style_monthly_sheet(ws, periods, len(output_rows))


# ====================================================================
#  INTERNAL – Monthly output row definitions
# ====================================================================


def _define_output_rows() -> List[Tuple[str, str | None, str | None]]:
    """Return the list of (label, unit, key) rows for the Monthly sheet."""
    return [
        # --- PRICE INDICATORS ---
        ("Price Indicators", None, None),
        ("EPEX forecast BL", "EUR/MWh", "EPEX forecast BL"),
        ("CDS", "EUR/MWh", "CDS"),
        ("", None, None),
        # --- REVENUES ---
        ("EPEX Revenues", "EUR", "EPEX Revenues"),
        ("DOW revenues (TOTAL)", "EUR", "DOW revenues (TOTAL)"),
        ("Total Revenues (TOTAL)", "EUR", "Total Revenues (TOTAL)"),
        ("", None, None),
        # --- COSTS ---
        ("    Coal API2 (TOTAL)", "EUR", "Coal API2"),
        ("        Merchant", "EUR", "Coal API2 -- Merchant"),
        ("        DOW", "EUR", "Coal API2 -- DOW"),
        ("    Coal other (TOTAL)", "EUR", "Coal other costs (premium, transportation)"),
        ("        Merchant", "EUR", "Coal other -- Merchant"),
        ("        DOW", "EUR", "Coal other -- DOW"),
        ("    CO2 costs (TOTAL)", "EUR", "CO2 costs"),
        ("        Merchant", "EUR", "CO2 costs -- Merchant"),
        ("        DOW", "EUR", "CO2 costs -- DOW"),
        ("Other costs (TOTAL)", "EUR", "Other costs (TOTAL)"),
        ("    Start costs", "EUR", "Start costs"),
        ("    Variable other costs", "EUR", "Variable other costs"),
        ("    House Power", "EUR", "House Power"),
        ("Total Costs (TOTAL)", "EUR", "Total Costs (TOTAL)"),
        ("PnL", "EUR", "PnL"),
        ("", None, None),
        # --- OPERATION STATISTICS ---
        ("Operation Statistics", None, None),
        ("Power produced Merchant (TOTAL)", "GWh", "Power produced Merchant"),
        ("Power produced DOW", "GWh", "Power produced DOW"),
        ("Coal Consumption TOTAL", "kMT", "Coal Consumption"),
        ("    Coal Consumption -- Merchant", "kMT", "Coal Consumption -- Merchant"),
        ("    Coal Consumption -- DOW", "kMT", "Coal Consumption -- DOW"),
        ("CO2 emissions TOTAL", "kMT", "CO2 emissions"),
        ("    CO2 emissions -- Merchant", "kMT", "CO2 emissions -- Merchant"),
        ("    CO2 emissions -- DOW", "kMT", "CO2 emissions -- DOW"),
        ("Number of starts", "#", "Number of starts"),
        ("Outage", "h", "Outage"),
        ("DOW volume", "GWhth", "DOW volume"),
        ("", None, None),
        # --- COAL CONSTRAINT SHADOW PRICES ---
        ("Coal Shadow Price", "EUR/t", "Coal shadow price"),
    ]


# ====================================================================
#  INTERNAL – Monthly Excel styling
# ====================================================================


def _style_monthly_sheet(ws, periods, num_output_rows: int) -> None:
    """Apply professional formatting to the Monthly worksheet."""
    col_start = 3  # column C
    last_col = col_start + len(periods) - 1 if periods else 2

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FFC00000")
    section_fill = PatternFill("solid", fgColor="FF714F")
    thin = Side(style="thin")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="top")
    thousands_fmt = "#,##0"

    # --- Row 1: merged year headers ---
    current_year = None
    year_start_col = col_start
    for j, (yy, mm) in enumerate(periods):
        col = col_start + j
        if current_year is None:
            current_year = yy
            year_start_col = col
        if yy != current_year:
            ws.merge_cells(
                start_row=1, start_column=year_start_col,
                end_row=1, end_column=col - 1,
            )
            ws.cell(row=1, column=year_start_col, value=str(current_year)).font = bold
            current_year = yy
            year_start_col = col
        ws.cell(row=2, column=col, value=cfg.MONTH_NAMES.get(mm, str(mm))).font = bold

    if current_year is not None:
        ws.merge_cells(
            start_row=1, start_column=year_start_col,
            end_row=1, end_column=last_col,
        )
        ws.cell(row=1, column=year_start_col, value=str(current_year)).font = bold

    # A1/A2/B2 blank
    for r in (1, 2):
        for c in (1, 2):
            ws.cell(row=r, column=c, value="")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    for j in range(len(periods)):
        ws.column_dimensions[get_column_letter(col_start + j)].width = 14 if j == 0 else 13

    ws.freeze_panes = "C4"

    # --- Helper functions ---
    def _style_row(row_idx, fill=None, bold_=False, border=None, align=None):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            if fill is not None:
                cell.fill = fill
            if bold_:
                cell.font = Font(bold=True)
            if border is not None:
                cell.border = border
            if align is not None:
                cell.alignment = align

    def _font_white(row_idx):
        for c in range(1, last_col + 1):
            ws.cell(row=row_idx, column=c).font = white_bold

    def _find_row(label):
        for rr in range(1, ws.max_row + 1):
            if ws.cell(row=rr, column=1).value == label:
                return rr
        return None

    # --- Header row (row 3) ---
    _style_row(3, fill=header_fill, bold_=True, border=header_border, align=header_align)
    _font_white(3)

    # --- Remove (TOTAL) text ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=last_col):
        for cell in row:
            if isinstance(cell.value, str) and "(TOTAL)" in cell.value:
                cell.value = re.sub(r"\s*\(TOTAL\)\s*", " ", cell.value).strip()
                cell.value = re.sub(r"\s{2,}", " ", cell.value).strip()

    # --- Style 'Price Indicators' section header ---
    pi_row = _find_row("Price Indicators")
    if pi_row is not None:
        _style_row(pi_row, fill=section_fill, bold_=True, border=header_border, align=header_align)
        _font_white(pi_row)

    # --- Insert 'Revenues' divider row above 'EPEX Revenues' ---
    mr_row = _find_row("EPEX Revenues")
    if mr_row is not None:
        ws.insert_rows(mr_row, amount=1)
        _style_row(mr_row, fill=section_fill, bold_=True, border=header_border, align=header_align)
        _font_white(mr_row)
        ws.cell(row=mr_row, column=1, value="Revenues")
        for c in range(2, last_col + 1):
            ws.cell(row=mr_row, column=c, value=None)

    # --- Style 'Total Revenues' ---
    tr_row = _find_row("Total Revenues")
    if tr_row is not None:
        _style_row(tr_row, fill=header_fill, bold_=False, border=header_border)
        _font_white(tr_row)

        # Insert 'Costs' divider
        ws.insert_rows(tr_row + 1, amount=1)
        costs_row = tr_row + 1
        _style_row(costs_row, fill=section_fill, bold_=True, border=header_border, align=header_align)
        _font_white(costs_row)
        ws.cell(row=costs_row, column=1, value="Costs")
        for c in range(2, last_col + 1):
            ws.cell(row=costs_row, column=c, value=None)

    # --- Style 'Total Costs' ---
    tc_row = _find_row("Total Costs")
    if tc_row is not None:
        _style_row(tc_row, fill=header_fill, bold_=False, border=header_border)
        _font_white(tc_row)

    # --- Style 'PnL' ---
    pnl_row = _find_row("PnL")
    if pnl_row is not None:
        _style_row(pnl_row, fill=section_fill, bold_=True, border=header_border)
        _font_white(pnl_row)

    # --- Section rows ---
    for r in range(4, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label in ("Income", "Total"):
            _style_row(r, fill=section_fill, bold_=True)

    # Operation Statistics header
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Operation Statistics":
            _style_row(r, fill=header_fill, bold_=True, border=header_border, align=header_align)
            _font_white(r)
            break

    # --- Number formatting ---
    for r in range(4, ws.max_row + 1):
        for c in range(3, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, numbers.Number):
                cell.number_format = thousands_fmt

    # Integer rows
    integer_rows = [
        "Number of starts", "Outage",
        "DOW volume",
    ]
    for label in integer_rows:
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == label:
                for c in range(3, last_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell.value, numbers.Number):
                        cell.value = int(round(cell.value))
                        cell.number_format = "0"
                break

    # EUR/MWh rows – two decimal places
    eur_mwh_rows = ["EPEX forecast BL", "CDS"]
    for label in eur_mwh_rows:
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == label:
                for c in range(3, last_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell.value, numbers.Number):
                        cell.number_format = "#,##0.00"
                break

    # Row 41 white font
    row_41 = 41
    if row_41 <= ws.max_row:
        for c in range(1, last_col + 1):
            ws.cell(row=row_41, column=c).font = Font(color="FFFFFF", bold=False)
